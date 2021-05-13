# code by William Russell

import openpyxl

class Card:

    numCards = 0
    numShiny = 0

    
    # define the initialiser
    def __init__(self, theName, theType, theHP, theMoves, isShiny):
        self.name = theName
        self.type = theType
        self.hp = theHP
        self.moves = theMoves
        self.shiny = isShiny

        # adding the normal and shiny cards as they are initialised
        Card.numCards += 1
        if(self.shiny == 1):
            Card.numShiny += 1

            
    # get moves of the card
    def getMoves(self, Move1, Damage1, Move2, Damage2, Move3, Damage3, Move4, Damage4, Move5, Damage5):

        # holding all these moves within a dictionary card_moves
        card_moves = {Move1: Damage1,
                      Move2: Damage2,
                      Move3: Damage3,
                      Move4: Damage4,
                      Move5: Damage5}

        # iterate through the items in the dictionary and appends these items into an empty list
        output_List = []
        for item in card_moves.items():
            output_List.append(list(item))

        # assigning the final variable from cards as a list of the moves
        self.moves = output_List

        
    # defining the shinyness of the card
    def getShiny(self):

        # returns true or false depending on whether the card is shiny or not
        if(self.shiny == 1):
            self.shiny = True
        elif(self.shiny == 0):
            self.shiny = False
        else:
            return

        
    # printing a description of the card
    def __str__(self):

        return f"""
        Name: {self.name}\n\n
        Type: {self.type}\n\n
        Health Points: {self.hp}\n\n
        Move 1: {self.moves[0][0]}\n
        Damage Factor 1: {self.moves[0][1]}\n\n
        Move 2: {self.moves[1][0]}\n
        Damage Factor 2: {self.moves[1][1]}\n\n
        Move 3: {self.moves[2][0]}\n
        Damage Factor 3: {self.moves[2][1]}\n\n
        Move 4: {self.moves[3][0]}\n
        Damage Factor 4: {self.moves[3][1]}\n\n
        Move 5: {self.moves[4][0]}\n
        Damage Factor 5: {self.moves[4][1]}\n\n
        Shiny: {self.shiny}\n\n
        """


# create class Deck

class Deck:

    
    def __init__(self):
        self.theDeck = []
        self.theShinyDeck = []

        
    def inputFromFile(self, theFileName):
        # opening up the excel file containing the cards
        try:
            book = openpyxl.load_workbook(theFileName)
            sheet = book.active
        except FileNotFoundError:
            print("\nThere is no such file available.")
            return

        # iterates through each row and adds the card into the deck

        Card_From_Input = []
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                Card_From_Input.append(cell.value)

            # to prevent the 'None' glitch in the program
            if(Card_From_Input[0] != None):

                # assign the cell value to the respective variables in class Card
                theName = Card_From_Input[0]
                theType = Card_From_Input[1]
                theHP = Card_From_Input[2]

                # assign the cell value of either 1 or 0 to the isShiny variable
                isShiny = Card_From_Input[3]

                # returns true or false depending on whether the card is shiny or not
                if(isShiny == 1):
                    isShiny = True
                elif(isShiny == 0):
                    isShiny = False

                # each move and their respective damage from the excel file
                Move1 = Card_From_Input[4]
                Damage1 = Card_From_Input[5]
                Move2 = Card_From_Input[6]
                Damage2 = Card_From_Input[7]
                Move3 = Card_From_Input[8]
                Damage3 = Card_From_Input[9]
                Move4 = Card_From_Input[10]
                Damage4 = Card_From_Input[11]
                Move5 = Card_From_Input[12]
                Damage5 = Card_From_Input[13]

                # holding all these moves within a dictionary card_moves
                card_moves = {Move1: Damage1,
                              Move2: Damage2,
                              Move3: Damage3,
                              Move4: Damage4,
                              Move5: Damage5}

                # iterate through the items in the dictionary and appends these items into an empty list
                output_List = []
                for item in card_moves.items():
                    output_List.append(list(item))

                # assigning the final variable from cards as a list of the moves
                theMoves = output_List

                # final instance template
                input_Card = Card(theName, theType, theHP, theMoves, isShiny)

                # to prevent the 'None' glitch in the program
                if(input_Card.name != None):
                    self.theDeck.append(input_Card)

                    # if the card is also shiny, it will be added to the shiny deck
                    if(isShiny == True):
                        self.theShinyDeck.append(input_Card)
                else:
                    return
            else:
                return

            # deletes the list containing the card to allow the loop to iterate over the next loop
            del Card_From_Input
            Card_From_Input = []

            
    # returns the summary of the deck
    def __str__(self):

        print("\nSummary of the Deck: ")

        # total number of cards in the deck
        print(f"\nNumber of Cards: {Card.numCards}\n")

        # total number of shiny cards in the deck
        print(f"Number of Shiny Cards: {Card.numShiny}")

        # calls the function getAverageDamage() to insert the average damage of the deck
        Deck.getAverageDamage(self)
        return

    
    # add a card to the deck
    def addCard(self, theCard):

        # Creating the type boundary, therefore the card can only be 1 of 6 values
        if(theCard.type in ["Fire", "Water", "Air", "Astral", "Earth", "Magi"]):

            # stating that the inputted values can only a binary number
            if(theCard.shiny == 0 or theCard.shiny == 1):
                self.theDeck.append(theCard)
                if(theCard.shiny == 1):

                    # if the card is also shiny, it will be added to the shiny deck
                    self.theShinyDeck.append(theCard)
            else:
                print("You must select a binary number!")
                return

        else:
            print("There is no such type!")
            return

        
    # remove a card from the deck
    def rmCard(self, theCard):
        try:
            self.theDeck.remove(theCard)
        except ValueError:

            return f"\nCannot remove {theCard.name}, as there is no such card in the deck.\n\n"

        
    # return the Card that is most powerful
    def getMostPowerful(self):

        # return only the cards that are present in the deck

        max_AVG = 0

        # loop through all the cards in the deck
        for i in range(len(self.theDeck)):

            # select the moves from the ith card
            powerMoves = self.theDeck[i].moves

            counter = 0
            total = 0

            # loop through the ith cards moves
            for j in range(0, 5):

                # to prevent the error when cards have moves that return 'none'
                if(powerMoves[j][1] != None):

                    # count the iterations
                    counter += 1

                    # add the damage factor of each move together
                    total += powerMoves[j][1]

            # formula to detect the average damage factor
            avDamageFactor = total/counter

            # method to find the highest average damage factor
            if(avDamageFactor >= max_AVG):
                max_AVG = avDamageFactor
                most_powerful_card = self.theDeck[i]

        # printing the most powerful card from the deck

        return f"\nThe most powerful card is:\n\n{most_powerful_card}\n\nMax Average Damage Factor: {max_AVG}\n"

    
    # return the average damage inflicted by all cards in the deck
    def getAverageDamage(self):

        # starting number of cards in the deck
        total_counter = 0

        # starting total of all the damage factors added together
        total_number = 0

        # loop through all the cards in the deck
        for i in range(len(self.theDeck)):

            # select the moves from the ith card
            powerMoves = self.theDeck[i].moves

            counter = 0
            total = 0

            # loop through the ith cards moves
            for j in range(0, 5):

                # to prevent the error when cards have moves that return 'none'
                if(powerMoves[j][1] != None):

                    # count the iterations
                    counter += 1

                    # add the damage factor of each move together
                    total += powerMoves[j][1]

            # formula to detect the average damage factor
            avDamageFactor = total/counter

            # increment the counter to count all the cards
            total_counter += 1

            # increment the total with the average damage factor of the ith card
            total_number += avDamageFactor

        # calculation to determine the average damage factor of all the cards
        averageDamage = round(total_number/total_counter, 1)
        print(f"\nAverage Damage: {averageDamage}\n\n")

        # printing the average damage factor to 1 decimal place
        return

    
    # print the information of all the cards in the deck
    def viewAllCards(self):

        print("\nView Cards:\n")

        # loop through all the cards
        for i in self.theDeck:
            print(i)

            
    # print the information of all the shiny cards in the deck
    def viewAllShinyCards(self):

        print("\nView Shiny Cards:\n")

        # loop through all the shiny cards
        for i in self.theShinyDeck:
            print(i)

            
    # print the information of all the cards in the deck that belong to the type of theType
    def viewAllByType(self, theType):

        print(f"\nType: {theType}\n")

        # loop through the all the cards
        for i in range(len(self.theDeck)):

            # if the type inserted into the instance matches the card type, then the information
            # of the card is printed
            cardType = self.theDeck[i].type
            if(theType == str(cardType)):
                print(self.theDeck[i])

                
    # return all cards held within the deck as a collection.
    def getCards(self):

        # empty list that will hold all the data for the cards in the collection
        getCardList = []

        for i in range(len(self.theDeck)):
            # get name
            name = self.theDeck[i].name
            getCardList.append(name)

            # get type
            type = self.theDeck[i].type
            getCardList.append(type)

            # get health points
            hp = self.theDeck[i].hp
            getCardList.append(hp)

            # get shiny
            shiny = self.theDeck[i].shiny
            if(shiny == True):
                shiny = 1
            elif(shiny == False):
                shiny = 0
            getCardList.append(shiny)

            # get moves
            for move in range(0, 5):
                getCardList.append(self.theDeck[i].moves[move][0])
                getCardList.append(self.theDeck[i].moves[move][1])

        return getCardList

    
    # This saves the Deck to an xlsx file
    def saveToFile(self, fileName):

        # creating an open workbook
        newBook = openpyxl.Workbook()
        newSheet = newBook.active

        # assigning a first row with headings
        firstRow = ("Name", "Type",	"HP", "Shiny", "Move Name 1", "Damage 1", "Move Name 2",
                    "Damage 2", "Move Name 3", "Damage 3", "Move Name 4", "Damage 4", "Move Name 5", "Damage 5")

        # adding the first row to the new sheet
        newSheet.append(firstRow)

        # placing all the card information into a list which is added as a row to the sheet, this is then looped
        # for all cards in the deck
        for i in range(len(self.theDeck)):

            rowList = []

            # get name
            name = self.theDeck[i].name
            rowList.append(name)

            # get type
            type = self.theDeck[i].type
            rowList.append(type)

            # get health points
            hp = self.theDeck[i].hp
            rowList.append(hp)

            # get shiny
            shiny = self.theDeck[i].shiny
            if(shiny == True):
                shiny = 1
            elif(shiny == False):
                shiny = 0
            rowList.append(shiny)

            # get moves
            for move in range(0, 5):
                rowList.append(self.theDeck[i].moves[move][0])
                rowList.append(self.theDeck[i].moves[move][1])

            newSheet.append(rowList)

        # saving the sheet to a file that the user determines
        newBook.save(fileName)

        # showing the user that a file has been created
        return f"\nThe file {fileName} has been created.\n"


# main body of code
def main():
    myDeck = Deck()

    # To import cards into the deck from a file
    myDeck.inputFromFile("sampleDeck.xlsx")

    # I have created some fake cards in order to demonstrate the functionality of my program
    c1 = Card("Jungi", "Fire", 230, "", 1)
    c2 = Card("Oogman", "Water", 200, "", 0)

    # To access the program, the user will also have to add the card moves
    c1.getMoves("Dragon Fist", 50, "Solum Blade", 50, "Iron Smack",
                67, "Tiger Spin", 76, "Laser Beam", 100)
    c2.getMoves("Water Dash", 54, "Laser Beam", 99, "Waterfall",
                73, "Whirlpool", 67, "Inferno Spring", 120)

    # To access the program, the user will also have to call the shiny function
    c1.getShiny()
    c2.getShiny()

    # To add these cards to the deck
    myDeck.addCard(c1)
    myDeck.addCard(c2)

    # Overall summary of the deck
    myDeck.__str__()

    # To view all cards in the deck
    myDeck.viewAllCards()

    # To view all the shiny cards in the deck
    myDeck.viewAllShinyCards()

    # To get the most powerful card
    myDeck.getMostPowerful()

    # To get the average damage of the deck
    myDeck.getAverageDamage()

    # To view the deck by a certain type
    myDeck.viewAllByType("Magi")

    # To remove cards from the deck (blurred out for now)
    # print(f"\nRemoving:\n\n{c1}\n")
    # myDeck.rmCard(c1)

    # To return the cards as a collection (list)
    print(myDeck.getCards())

    # Save your deck to a new file
    print(myDeck.saveToFile("firstDeck.xlsx"))


main()
